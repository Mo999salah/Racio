"""Generate synthetic Phase 7 XLSX fixtures. No real financial data is used."""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.utils.datetime import CALENDAR_MAC_1904

ROOT = Path(__file__).resolve().parent


def save(name: str, workbook: Workbook) -> Path:
    path = ROOT / name
    workbook.save(path)
    return path


def rewrite_archive(source: Path, target: Path, replacements: dict[str, bytes]) -> None:
    with ZipFile(source) as original, ZipFile(target, "w", ZIP_DEFLATED) as output:
        for info in original.infolist():
            output.writestr(info.filename, replacements.get(info.filename, original.read(info)))
        for name, payload in replacements.items():
            if name not in original.namelist():
                output.writestr(name, payload)


def english_statement() -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Statement"
    sheet.append(["Date", "Description", "Amount", "Currency", "Balance"])
    sheet.append([date(2026, 1, 2), "Neighbourhood market", -12.34, "USD", 987.66])
    sheet.append([date(2026, 1, 3), "Salary", 1000, "USD", 1987.66])
    sheet.append([])
    sheet.append([date(2026, 1, 3), "Salary", 1000, "USD", 1987.66])
    sheet.append(["", "Total", 1987.66, "USD", 1987.66])
    sheet["A2"].number_format = "yyyy-mm-dd"
    sheet["A3"].number_format = "yyyy-mm-dd"
    sheet["A5"].number_format = "yyyy-mm-dd"
    sheet["C2"].number_format = "0.00"
    sheet["E2"].number_format = "0.00"
    return save("english-one-sheet.xlsx", workbook)


def turkish_statement() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ekstre"
    sheet.append(["Tarih", "Açıklama", "Tutar", "Para Birimi", "Bakiye"])
    sheet.append([date(2026, 2, 1), "Yerel market", -45.9, "TRY", 954.1])
    sheet["A2"].number_format = "dd.mm.yyyy"
    sheet["C2"].number_format = '#,##0.00 [$₺-tr-TR]'
    sheet["E2"].number_format = '#,##0.00 [$₺-tr-TR]'
    save("turkish-decimal-comma.xlsx", workbook)


def arabic_statement() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "كشف الحساب"
    sheet.sheet_view.rightToLeft = True
    sheet.append(["التاريخ", "البيان", "مدين", "دائن", "العملة", "الرصيد"])
    sheet.append(["01/03/2026", "متجر الحي", "25.500", "", "AED", "974.500"])
    sheet.append(["02/03/2026", "إيداع", "", "100.000000", "AED", "1074.500"])
    save("arabic-debit-credit.xlsx", workbook)


def multiple_and_hidden() -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "January"
    first.append(["Date", "Description", "Amount", "Currency"])
    first.append(["2026-01-01", "Market", "-10.00", "EUR"])
    second = workbook.create_sheet("February")
    second.append(["Date", "Description", "Amount", "Currency"])
    second.append(["2026-02-01", "Salary", "800.00", "EUR"])
    hidden = workbook.create_sheet("Archive")
    hidden.append(["Date", "Description", "Amount", "Currency"])
    hidden.append(["2025-12-01", "Archived row", "1.00", "EUR"])
    hidden.sheet_state = "hidden"
    very_hidden = workbook.create_sheet("Internal")
    very_hidden["A1"] = "Not importable"
    very_hidden.sheet_state = "veryHidden"
    save("multiple-visible-and-hidden.xlsx", workbook)


def title_merged_and_repeated() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.merge_cells("A1:E1")
    sheet["A1"] = "Synthetic statement January 2026"
    sheet.append([])
    headers = ["Date", "Description", "Amount", "Currency", "Balance"]
    sheet.append(headers)
    sheet.append(["2026-01-02", "Market", "-12.340000", "USD", "87.660000"])
    sheet.append(headers)
    sheet.append(["2026-01-03", "Salary", "100.123456", "USD", "187.783456"])
    sheet.append(["", "Closing balance", "", "USD", "187.783456"])
    save("title-merged-repeated-footer.xlsx", workbook)


def dates_and_text_amounts() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mixed values"
    sheet.append(["Date", "Description", "Amount", "Currency"])
    sheet.append([date(2026, 4, 1), "Date cell", "12.34", "GBP"])
    sheet.append(["02.04.2026", "Text date", "12.345", "GBP"])
    sheet.append(["03/04/2026", "Ambiguous text date", "0.123456", "GBP"])
    sheet.append(["2026-04-04", "Invalid precision", "0.1234567", "GBP"])
    sheet.append(["2026-04-05", "Multiple currencies", "5", "EUR"])
    sheet["A2"].number_format = "yyyy-mm-dd"
    save("dates-text-amounts-multiple-currencies.xlsx", workbook)


def date_1904() -> None:
    workbook = Workbook()
    workbook.epoch = CALENDAR_MAC_1904
    sheet = workbook.active
    sheet.title = "1904 dates"
    sheet.append(["Date", "Description", "Amount", "Currency"])
    sheet.append([date(2026, 5, 1), "1904 date system", -1, "USD"])
    sheet["A2"].number_format = "yyyy-mm-dd"
    save("date-system-1904.xlsx", workbook)


def formulas() -> tuple[Path, Path]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Formula values"
    sheet.append(["Date", "Description", "Amount", "Currency"])
    sheet.append(["2026-06-01", "Formula amount", "=12.34", "USD"])
    uncached = save("formula-without-cache.xlsx", workbook)
    with ZipFile(uncached) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml")
    cached_xml = re.sub(
        br"(<f>[^<]+</f>)<v\s*/>",
        br"\g<1><v>12.34</v>",
        xml,
    )
    cached = ROOT / "formula-with-cached-value.xlsx"
    rewrite_archive(uncached, cached, {"xl/worksheets/sheet1.xml": cached_xml})
    return uncached, cached


def excessive_dimensions() -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Too many rows")
    sheet.append(["Date", "Description", "Amount", "Currency"])
    for row in range(100_001):
        sheet.append(["2026-01-01", f"Synthetic {row}", "1.00", "USD"])
    save("excessive-rows.xlsx", workbook)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Too many columns"
    for column in range(1, 258):
        sheet.cell(1, column, f"Column {column}")
    save("excessive-columns.xlsx", workbook)


def unsafe_archives(valid: Path) -> None:
    with ZipFile(valid) as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        rels = archive.read("xl/_rels/workbook.xml.rels")
        content_types = archive.read("[Content_Types].xml")
    rewrite_archive(valid, ROOT / "zip-traversal.xlsx", {"../outside.xml": b"unsafe"})
    rewrite_archive(
        valid,
        ROOT / "compression-bomb-simulation.xlsx",
        {"xl/media/highly-compressible.bin": b"0" * (2 * 1024 * 1024)},
    )
    rewrite_archive(
        valid,
        ROOT / "malformed-workbook.xlsx",
        {"xl/workbook.xml": workbook_xml[:100] + b"<broken"},
    )
    external = rels.replace(
        b"</Relationships>",
        b'<Relationship Id="external" Type="external" Target="https://example.invalid" '
        b'TargetMode="External"/></Relationships>',
    )
    rewrite_archive(
        valid,
        ROOT / "external-link.xlsx",
        {"xl/_rels/workbook.xml.rels": external},
    )
    missing_target = rels.replace(
        b"worksheets/sheet1.xml",
        b"worksheets/missing-sheet.xml",
    )
    rewrite_archive(
        valid,
        ROOT / "missing-relationship-target.xlsx",
        {"xl/_rels/workbook.xml.rels": missing_target},
    )
    macro_types = content_types.replace(
        b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    rewrite_archive(
        valid,
        ROOT / "macro-enabled.xlsm",
        {"[Content_Types].xml": macro_types, "xl/vbaProject.bin": b"synthetic"},
    )
    (ROOT / "fake-binary.xlsx").write_bytes(b"not an OOXML workbook")
    (ROOT / "unsupported-legacy.xls").write_bytes(
        b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1synthetic legacy marker"
    )


def floating_point_artifact(valid: Path) -> None:
    with ZipFile(valid) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml")
    artifact = xml.replace(b"<v>-12.34</v>", b"<v>-12.340000000000002</v>", 1)
    rewrite_archive(
        valid,
        ROOT / "floating-point-artifact.xlsx",
        {"xl/worksheets/sheet1.xml": artifact},
    )


def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)
    valid = english_statement()
    turkish_statement()
    arabic_statement()
    multiple_and_hidden()
    title_merged_and_repeated()
    dates_and_text_amounts()
    date_1904()
    formulas()
    floating_point_artifact(valid)
    excessive_dimensions()
    unsafe_archives(valid)


if __name__ == "__main__":
    main()
