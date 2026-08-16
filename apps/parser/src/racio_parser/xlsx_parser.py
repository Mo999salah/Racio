from __future__ import annotations

import io
import posixpath
import re
import zipfile
from datetime import date, datetime, timedelta
from decimal import ROUND_FLOOR, Decimal, InvalidOperation
from typing import Any, Literal, cast
from xml.etree import ElementTree

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell import Cell  # type: ignore[import-untyped]
from openpyxl.utils import (  # type: ignore[import-untyped]
    column_index_from_string,
    get_column_letter,
)

from .config import ParserSettings
from .csv_parser import _clean, _decimal, _parse_date, infer_mapping
from .models import (
    RawWorkbookCell,
    WorkbookInspection,
    WorkbookSheetInspection,
    XlsxMapping,
    XlsxMappingResult,
    XlsxParsedCandidate,
    XlsxParserResult,
    XlsxParserSource,
)
from .xlsx_security import XlsxSecurityError, validate_xlsx_archive

SHEET_TAG = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
RELATIONSHIP_TAG = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
RELATIONSHIP_ID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
CELL_REFERENCE = re.compile(r"^([A-Z]{1,3})([1-9]\d*)$")
SUMMARY_WORDS = {
    "total",
    "totals",
    "summary",
    "closingbalance",
    "toplam",
    "geneltoplam",
    "kapanışbakiyesi",
    "الاجمالي",
    "الإجمالي",
    "الرصيدالختامي",
}
MAPPED_FIELDS = (
    "bookingDate",
    "valueDate",
    "description",
    "amount",
    "debit",
    "credit",
    "currency",
    "balance",
    "counterparty",
    "transactionIdentifier",
)


def _safe_xml(payload: bytes) -> ElementTree.Element:
    prefix = payload[:8_192].upper()
    if b"<!DOCTYPE" in prefix or b"<!ENTITY" in prefix:
        raise XlsxSecurityError("XLSX_INVALID_XML")
    try:
        return ElementTree.fromstring(payload)
    except ElementTree.ParseError as error:
        raise XlsxSecurityError("XLSX_INVALID_XML") from error


def _sheet_parts(content: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        workbook = _safe_xml(archive.read("xl/workbook.xml"))
        relationships = _safe_xml(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            node.attrib.get("Id", ""): node.attrib.get("Target", "")
            for node in relationships.iter(RELATIONSHIP_TAG)
        }
        parts: list[str] = []
        for sheet in workbook.iter(SHEET_TAG):
            relationship_id = sheet.attrib.get(RELATIONSHIP_ID, "")
            target = targets.get(relationship_id, "")
            if not target:
                raise XlsxSecurityError("XLSX_INVALID_WORKBOOK")
            if target.startswith("/"):
                part = target.lstrip("/")
            else:
                part = posixpath.normpath(posixpath.join("xl", target))
            if part.startswith("../") or part not in archive.namelist():
                raise XlsxSecurityError("XLSX_INVALID_WORKBOOK")
            parts.append(part)
        return parts


def _inspect_sheet_xml(
    content: bytes, part: str, settings: ParserSettings
) -> tuple[int, int, int, int, int]:
    max_row = 0
    max_column = 0
    populated = 0
    formulas = 0
    merged = 0
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive, archive.open(part) as stream:
            prefix = stream.read(8_192).upper()
            if b"<!DOCTYPE" in prefix or b"<!ENTITY" in prefix:
                raise XlsxSecurityError("XLSX_INVALID_XML")
            stream.seek(0)
            for _event, node in ElementTree.iterparse(stream, events=("end",)):
                local_name = node.tag.rsplit("}", 1)[-1]
                if local_name == "c":
                    reference = node.attrib.get("r", "")
                    match = CELL_REFERENCE.fullmatch(reference)
                    if not match:
                        raise XlsxSecurityError("XLSX_INVALID_XML")
                    column = column_index_from_string(match.group(1))
                    row = int(match.group(2))
                    has_value = any(
                        child.tag.rsplit("}", 1)[-1] in {"v", "is", "f"}
                        and (child.text is not None or len(child))
                        for child in node
                    )
                    if has_value:
                        populated += 1
                        max_row = max(max_row, row)
                        max_column = max(max_column, column)
                        if len("".join(node.itertext())) > settings.max_xlsx_cell_string_length:
                            raise XlsxSecurityError("XLSX_CELL_STRING_LIMIT_EXCEEDED")
                    if any(child.tag.rsplit("}", 1)[-1] == "f" for child in node):
                        formulas += 1
                    node.clear()
                elif local_name == "mergeCell":
                    merged += 1
                    node.clear()
                if max_row > settings.max_xlsx_rows:
                    raise XlsxSecurityError("XLSX_ROW_LIMIT_EXCEEDED")
                if max_column > settings.max_xlsx_columns:
                    raise XlsxSecurityError("XLSX_COLUMN_LIMIT_EXCEEDED")
                if populated > settings.max_xlsx_populated_cells:
                    raise XlsxSecurityError("XLSX_CELL_LIMIT_EXCEEDED")
                if formulas > settings.max_xlsx_formulas:
                    raise XlsxSecurityError("XLSX_FORMULA_LIMIT_EXCEEDED")
                if merged > settings.max_xlsx_merged_ranges:
                    raise XlsxSecurityError("XLSX_MERGED_RANGE_LIMIT_EXCEEDED")
    except ElementTree.ParseError as error:
        raise XlsxSecurityError("XLSX_INVALID_XML") from error
    return max_row, max_column, populated, merged, formulas


def _date_system(workbook: Any) -> Literal["1900", "1904"]:
    return "1904" if getattr(workbook.epoch, "year", 1899) == 1904 else "1900"


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return _clean(str(value))


def inspect_xlsx(content: bytes, filename: str, settings: ParserSettings) -> WorkbookInspection:
    del filename
    validate_xlsx_archive(content, settings)
    parts = _sheet_parts(content)
    if len(parts) > settings.max_xlsx_sheets:
        raise XlsxSecurityError("XLSX_SHEET_LIMIT_EXCEEDED")
    sheet_statistics = [_inspect_sheet_xml(content, part, settings) for part in parts]
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=False, keep_links=False
        )
        cached_workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
    except Exception as error:
        raise XlsxSecurityError("XLSX_INVALID_WORKBOOK") from error
    try:
        if len(workbook.worksheets) != len(parts):
            raise XlsxSecurityError("XLSX_INVALID_WORKBOOK")
        sheets: list[WorkbookSheetInspection] = []
        for index, (worksheet, cached_worksheet, statistics) in enumerate(
            zip(
                workbook.worksheets,
                cached_workbook.worksheets,
                sheet_statistics,
                strict=True,
            )
        ):
            rows, columns, populated, merged, formulas = statistics
            sample_rows: list[list[str]] = []
            formula_rows = worksheet.iter_rows(
                min_row=1, max_row=min(rows, 8), max_col=min(columns, 16)
            )
            cached_rows = cached_worksheet.iter_rows(
                min_row=1, max_row=min(rows, 8), max_col=min(columns, 16)
            )
            for formula_row, cached_row in zip(formula_rows, cached_rows, strict=True):
                sample: list[str] = []
                for formula_cell, cached_cell in zip(formula_row, cached_row, strict=True):
                    value = (
                        cached_cell.value if formula_cell.data_type == "f" else formula_cell.value
                    )
                    sample.append(_display_value(value)[:2_000])
                sample_rows.append(sample)
            warnings: list[str] = []
            if worksheet.sheet_state == "hidden":
                warnings.append("hidden_sheet")
            if worksheet.sheet_state == "veryHidden":
                warnings.append("very_hidden_sheet")
            if formulas:
                warnings.append("formulas_present")
            if merged:
                warnings.append("merged_cells_present")
            if populated == 0:
                warnings.append("empty_sheet")
            sheets.append(
                WorkbookSheetInspection(
                    id=f"sheet-{index}",
                    name=worksheet.title,
                    index=index,
                    hidden=worksheet.sheet_state != "visible",
                    veryHidden=worksheet.sheet_state == "veryHidden",
                    estimatedRows=rows,
                    estimatedColumns=columns,
                    populatedCells=populated,
                    mergedRangeCount=merged,
                    formulaCellCount=formulas,
                    sampleRows=sample_rows,
                    warnings=warnings,
                )
            )
        if not any(not sheet.veryHidden and sheet.populatedCells for sheet in sheets):
            raise XlsxSecurityError("XLSX_NO_USABLE_SHEET")
        return WorkbookInspection(
            contractVersion="racio.workbook-inspection.v1",
            workbookType="xlsx",
            sheetCount=len(sheets),
            dateSystem=_date_system(workbook),
            sheets=sheets,
            workbookWarnings=[],
        )
    finally:
        workbook.close()
        cached_workbook.close()


def _choose_header(
    worksheet: Any, cached_worksheet: Any, max_columns: int
) -> tuple[list[str], Any, str, float, list[str], int]:
    candidates: list[tuple[int, int, list[str], Any, str, float, list[str], int]] = []
    formula_rows = worksheet.iter_rows(min_row=1, max_row=50, max_col=max_columns)
    cached_rows = cached_worksheet.iter_rows(min_row=1, max_row=50, max_col=max_columns)
    for row_number, (formula_row, cached_row) in enumerate(
        zip(formula_rows, cached_rows, strict=True), start=1
    ):
        values = [
            _display_value(cached.value if original.data_type == "f" else original.value)
            for original, cached in zip(formula_row, cached_row, strict=True)
        ]
        if not any(values):
            continue
        mapping, status, confidence, warnings = infer_mapping(values)
        mapped = sum(getattr(mapping, field) is not None for field in MAPPED_FIELDS)
        required = sum(
            [
                mapping.bookingDate is not None,
                mapping.description is not None,
                mapping.amount is not None
                or mapping.debit is not None
                or mapping.credit is not None,
            ]
        )
        candidates.append(
            (required, mapped, values, mapping, status, confidence, warnings, row_number)
        )
    if not candidates:
        empty_mapping, _, _, warnings = infer_mapping([])
        return [], empty_mapping, "invalid", 0.0, [*warnings, "header_not_found"], 1
    candidates.sort(key=lambda item: (item[0], item[1], -item[7]), reverse=True)
    best = candidates[0]
    tied = len(candidates) > 1 and candidates[1][:2] == best[:2]
    repeated_same_header = tied and [_clean(value).casefold() for value in candidates[1][2]] == [
        _clean(value).casefold() for value in best[2]
    ]
    ambiguous_tie = tied and not repeated_same_header
    status = "ambiguous" if ambiguous_tie or best[0] < 3 or best[4] != "confident" else "confident"
    warnings = list(best[6])
    if ambiguous_tie:
        warnings.append("ambiguous_header")
    return best[2], best[3], status, best[5] if status == "confident" else 0.55, warnings, best[7]


def _raw_cell_tokens(
    content: bytes,
    part: str,
    selected_columns: set[int],
    first_row: int,
    last_row: int,
) -> dict[str, tuple[str | None, str | None, str | None]]:
    tokens: dict[str, tuple[str | None, str | None, str | None]] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive, archive.open(part) as stream:
            for _event, node in ElementTree.iterparse(stream, events=("end",)):
                if node.tag.rsplit("}", 1)[-1] != "c":
                    continue
                coordinate = node.attrib.get("r", "")
                match = CELL_REFERENCE.fullmatch(coordinate)
                if not match:
                    raise XlsxSecurityError("XLSX_INVALID_XML")
                column = column_index_from_string(match.group(1)) - 1
                row = int(match.group(2))
                if column in selected_columns and first_row <= row <= last_row:
                    value: str | None = None
                    formula: str | None = None
                    for child in node:
                        local_name = child.tag.rsplit("}", 1)[-1]
                        if local_name == "v":
                            value = child.text
                        elif local_name == "f":
                            formula = child.text
                    tokens[coordinate] = (value, formula, node.attrib.get("t"))
                node.clear()
    except ElementTree.ParseError as error:
        raise XlsxSecurityError("XLSX_INVALID_XML") from error
    return tokens


def _fixed_decimal_places(number_format: str) -> int | None:
    section = number_format.split(";", 1)[0]
    section = re.sub(r'"[^"]*"|\\.|_.|\[[^\]]*\]', "", section)
    if "E" in section.upper() or "." not in section:
        return 0 if any(character in section for character in "0#") else None
    fraction = section.split(".", 1)[1]
    match = re.match(r"([0#]+)", fraction)
    return len(match.group(1)) if match else 0


def _decimal_from_numeric(raw: str, number_format: str) -> tuple[str | None, bool]:
    try:
        number = Decimal(raw)
    except InvalidOperation:
        return None, False
    if not number.is_finite():
        return None, False
    precision_normalized = False
    sign = -1 if number < 0 else 1
    number = abs(number)
    formatted = format(number, "f")
    whole, _, fraction = formatted.partition(".")
    if len(whole.lstrip("0") or "0") > 14:
        return None, False
    if len(fraction.rstrip("0")) > 6:
        places = _fixed_decimal_places(number_format)
        if places is None or places > 6:
            return None, False
        quantum = Decimal(1).scaleb(-places)
        normalized = number.quantize(quantum)
        if abs(number - normalized) > Decimal("0.000000000001"):
            return None, False
        number = normalized
        precision_normalized = True
    result = format(number, "f")
    result = result.rstrip("0").rstrip(".") if "." in result else result
    if sign < 0 and result != "0":
        result = f"-{result}"
    return result, precision_normalized


def _excel_serial_date(raw: str, date_system: str) -> str | None:
    try:
        serial = Decimal(raw)
    except InvalidOperation:
        return None
    if not serial.is_finite() or serial < 0:
        return None
    whole_days = int(serial.to_integral_value(rounding=ROUND_FLOOR))
    if date_system == "1900":
        if whole_days == 60:
            return None
        adjusted = whole_days if whole_days < 60 else whole_days - 1
        base = date(1899, 12, 31)
    else:
        adjusted = whole_days
        base = date(1904, 1, 1)
    try:
        return (base + timedelta(days=adjusted)).isoformat()
    except (OverflowError, ValueError):
        return None


def _parse_text_date(value: str, date_format: str | None) -> tuple[str | None, bool]:
    normalized = _clean(value)
    if not normalized:
        return None, False
    if date_format:
        return _parse_date(normalized, date_format), False
    ambiguous = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})[./-](\d{4})", normalized)
    if ambiguous:
        first, second = int(ambiguous.group(1)), int(ambiguous.group(2))
        if first <= 12 and second <= 12 and first != second:
            return None, True
    return _parse_date(normalized, None), False


def _cell_date(
    cell: Cell, raw_token: str | None, date_system: str, date_format: str | None
) -> tuple[str | None, bool]:
    if isinstance(cell.value, datetime):
        return cell.value.date().isoformat(), False
    if isinstance(cell.value, date):
        return cell.value.isoformat(), False
    if raw_token is not None and cell.is_date:
        return _excel_serial_date(raw_token, date_system), False
    if isinstance(cell.value, str):
        return _parse_text_date(cell.value, date_format)
    return None, False


def _mapped_cell(row: tuple[Cell, ...], index: int | None) -> Cell | None:
    return row[index] if index is not None and index < len(row) else None


def _cell_text(cell: Cell | None) -> str:
    return _display_value(cell.value) if cell is not None else ""


def _raw_cell(
    original: Cell,
    effective: Cell,
    raw_token: str | None,
    formula: str | None,
) -> RawWorkbookCell:
    has_cached = original.data_type == "f" and effective.value is not None
    if original.data_type == "f":
        raw_type = "formula_cached" if has_cached else "formula_uncached"
    elif original.data_type == "e":
        raw_type = "error"
    elif effective.value is None:
        raw_type = "blank"
    elif isinstance(effective.value, bool):
        raw_type = "boolean"
    elif isinstance(effective.value, (date, datetime)) or effective.is_date:
        raw_type = "date"
    elif effective.data_type == "n":
        raw_type = "number"
    else:
        raw_type = "string"
    return RawWorkbookCell(
        row=original.row,
        column=original.column,
        coordinate=original.coordinate,
        displayedText=_display_value(effective.value) or None,
        rawType=cast(
            Literal[
                "blank",
                "string",
                "number",
                "date",
                "boolean",
                "formula_cached",
                "formula_uncached",
                "error",
            ],
            raw_type,
        ),
        rawValue=raw_token if raw_token is not None else (_display_value(effective.value) or None),
        numberFormat=original.number_format or None,
        formula=(str(original.value)[:2_000] if formula is not None else None),
        hasCachedValue=has_cached if original.data_type == "f" else None,
    )


def _mapping_with_source(
    inferred: Any,
    sheet_index: int,
    sheet_name: str,
    header_row: int,
    first_data_row: int,
    last_data_row: int,
) -> XlsxMapping:
    values = inferred.model_dump()
    values.pop("headerRow", None)
    column_letters = {
        field: get_column_letter(index + 1)
        for field in MAPPED_FIELDS
        if (index := values.get(field)) is not None
    }
    return XlsxMapping(
        **values,
        sourceType="xlsx",
        selectedSheetId=f"sheet-{sheet_index}",
        selectedSheetName=sheet_name,
        selectedSheetIndex=sheet_index,
        headerRow=header_row,
        firstDataRow=first_data_row,
        lastDataRow=last_data_row,
        columnLetters=column_letters,
    )


def _effective_cells(
    formula_row: tuple[Cell, ...], cached_row: tuple[Cell, ...]
) -> tuple[Cell, ...]:
    return tuple(
        cached if original.data_type == "f" else original
        for original, cached in zip(formula_row, cached_row, strict=True)
    )


def parse_xlsx(
    content: bytes,
    filename: str,
    media_type: str,
    sheet_index: int,
    settings: ParserSettings,
    mapping_override: dict[str, Any] | None = None,
) -> XlsxParserResult:
    inspection = inspect_xlsx(content, filename, settings)
    if sheet_index < 0 or sheet_index >= len(inspection.sheets):
        raise XlsxSecurityError("XLSX_STALE_SHEET_SELECTION")
    sheet_info = inspection.sheets[sheet_index]
    if sheet_info.veryHidden:
        raise XlsxSecurityError("XLSX_NO_USABLE_SHEET")
    parts = _sheet_parts(content)
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=False, keep_links=False
        )
        cached_workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
    except Exception as error:
        raise XlsxSecurityError("XLSX_INVALID_WORKBOOK") from error
    try:
        worksheet = workbook.worksheets[sheet_index]
        cached_worksheet = cached_workbook.worksheets[sheet_index]
        if worksheet.title != sheet_info.name:
            raise XlsxSecurityError("XLSX_STALE_SHEET_SELECTION")
        mapping_warnings: list[str]
        if mapping_override:
            mapping = XlsxMapping.model_validate(mapping_override)
            if (
                mapping.selectedSheetIndex != sheet_index
                or mapping.selectedSheetName != worksheet.title
                or mapping.selectedSheetId != f"sheet-{sheet_index}"
            ):
                raise XlsxSecurityError("XLSX_STALE_SHEET_SELECTION")
            mapping_status, mapping_confidence, mapping_warnings = "confident", 1.0, []
            headers: list[str] = []
            header_rows = worksheet.iter_rows(
                min_row=mapping.headerRow,
                max_row=mapping.headerRow,
                max_col=sheet_info.estimatedColumns,
            )
            cached_header_rows = cached_worksheet.iter_rows(
                min_row=mapping.headerRow,
                max_row=mapping.headerRow,
                max_col=sheet_info.estimatedColumns,
            )
            for original_row, cached_row in zip(header_rows, cached_header_rows, strict=True):
                headers = [
                    _display_value(effective.value)
                    for effective in _effective_cells(original_row, cached_row)
                ]
        else:
            (
                headers,
                inferred,
                mapping_status,
                mapping_confidence,
                mapping_warnings,
                header_row,
            ) = _choose_header(worksheet, cached_worksheet, max(1, sheet_info.estimatedColumns))
            mapping = _mapping_with_source(
                inferred,
                sheet_index,
                worksheet.title,
                header_row,
                header_row + 1,
                max(header_row + 1, sheet_info.estimatedRows),
            )

        max_column_index = max(
            (
                getattr(mapping, field)
                for field in MAPPED_FIELDS
                if getattr(mapping, field) is not None
            ),
            default=0,
        )
        if max_column_index + 1 > settings.max_xlsx_columns:
            raise XlsxSecurityError("XLSX_COLUMN_LIMIT_EXCEEDED")
        last_row = mapping.lastDataRow or sheet_info.estimatedRows
        if last_row > settings.max_xlsx_rows or mapping.firstDataRow > last_row:
            raise XlsxSecurityError("XLSX_ROW_LIMIT_EXCEEDED")
        selected_columns = {
            index for field in MAPPED_FIELDS if (index := getattr(mapping, field)) is not None
        }
        raw_tokens = _raw_cell_tokens(
            content,
            parts[sheet_index],
            selected_columns,
            mapping.firstDataRow,
            last_row,
        )
        formula_rows = worksheet.iter_rows(
            min_row=mapping.firstDataRow,
            max_row=last_row,
            max_col=max_column_index + 1,
        )
        cached_rows = cached_worksheet.iter_rows(
            min_row=mapping.firstDataRow,
            max_row=last_row,
            max_col=max_column_index + 1,
        )
        candidates: list[XlsxParsedCandidate] = []
        parser_warnings = list(mapping_warnings)
        normalized_headers = [_clean(value).casefold() for value in headers]
        for original_row, cached_row in zip(formula_rows, cached_rows, strict=True):
            effective_row = _effective_cells(original_row, cached_row)
            visible_values = [_cell_text(cell) for cell in effective_row]
            if not any(visible_values):
                continue
            if (
                normalized_headers
                and [
                    _clean(value).casefold() for value in visible_values[: len(normalized_headers)]
                ]
                == normalized_headers
            ):
                if "repeated_header_row" not in parser_warnings:
                    parser_warnings.append("repeated_header_row")
                continue

            warnings: list[str] = []
            raw_cells: list[RawWorkbookCell] = []
            for column_index in sorted(selected_columns):
                original = original_row[column_index]
                effective = effective_row[column_index]
                raw_value, formula, _cell_type = raw_tokens.get(
                    original.coordinate, (None, None, None)
                )
                raw_cells.append(_raw_cell(original, effective, raw_value, formula))
                if original.data_type == "f":
                    warnings.append(
                        "formula_cached_value"
                        if effective.value is not None
                        else "formula_value_unavailable"
                    )

            def cells(
                field: str,
                source_row: tuple[Cell, ...] = original_row,
                value_row: tuple[Cell, ...] = effective_row,
            ) -> tuple[Cell | None, Cell | None, str | None]:
                column_index = getattr(mapping, field)
                original = _mapped_cell(source_row, column_index)
                effective = _mapped_cell(value_row, column_index)
                raw = (
                    raw_tokens.get(original.coordinate, (None, None, None))[0]
                    if original is not None
                    else None
                )
                return original, effective, raw

            booking_original, booking_cell, booking_raw_token = cells("bookingDate")
            value_original, value_cell, value_raw_token = cells("valueDate")
            booking_date, booking_ambiguous = (
                _cell_date(
                    booking_cell, booking_raw_token, inspection.dateSystem, mapping.dateFormat
                )
                if booking_cell is not None
                else (None, False)
            )
            value_date, value_ambiguous = (
                _cell_date(value_cell, value_raw_token, inspection.dateSystem, mapping.dateFormat)
                if value_cell is not None
                else (None, False)
            )
            booking_raw = _cell_text(booking_cell) or None
            value_raw = _cell_text(value_cell) or None
            if booking_raw and booking_date is None:
                warnings.append(
                    "ambiguous_booking_date" if booking_ambiguous else "invalid_booking_date"
                )
            if value_raw and value_date is None:
                warnings.append("ambiguous_value_date" if value_ambiguous else "invalid_value_date")
            if (
                booking_original is not None
                and booking_original.data_type == "f"
                and booking_cell is not None
                and booking_cell.value is None
            ):
                booking_date = None

            def amount_value(
                field: str,
                row_warnings: list[str] = warnings,
            ) -> tuple[str, str | None]:
                original, effective, raw = cells(field)
                if effective is None:
                    return "", None
                displayed = _cell_text(effective)
                if original is not None and original.data_type == "f" and effective.value is None:
                    return displayed, None
                if raw is not None and (
                    effective.data_type == "n" or isinstance(effective.value, (int, float, Decimal))
                ):
                    parsed, normalized = _decimal_from_numeric(
                        raw, original.number_format if original is not None else ""
                    )
                    if normalized:
                        row_warnings.append("precision_normalized_from_display_format")
                    return displayed or raw, parsed
                return displayed, _decimal(
                    displayed, mapping.decimalSeparator, mapping.thousandsSeparator
                )

            raw_amount, amount = amount_value("amount")
            raw_debit, debit = amount_value("debit")
            raw_credit, credit = amount_value("credit")
            direction = "unknown"
            if credit is not None and debit is None:
                amount, direction = credit.lstrip("-"), "credit"
            elif debit is not None and credit is None:
                amount, direction = debit.lstrip("-"), "debit"
            elif amount is not None:
                direction = "debit" if amount.startswith("-") else "credit"
                amount = amount.lstrip("-")
            if amount is None:
                warnings.append("invalid_amount")
            if direction == "unknown":
                warnings.append("unknown_direction")

            _description_original, description_cell, _description_raw = cells("description")
            description = _cell_text(description_cell)
            if not description:
                warnings.append("missing_description")
            _currency_original, currency_cell, _currency_raw = cells("currency")
            raw_currency = _cell_text(currency_cell).upper() or None
            currency = (
                raw_currency
                if raw_currency is not None and re.fullmatch(r"[A-Z]{3}", raw_currency)
                else None
            )
            if raw_currency and currency is None:
                warnings.append("invalid_currency")
            raw_balance, balance = amount_value("balance")
            _counterparty_original, counterparty_cell, _counterparty_raw = cells("counterparty")
            _transaction_original, transaction_cell, _transaction_raw = cells(
                "transactionIdentifier"
            )
            counterparty = _cell_text(counterparty_cell) or None
            transaction_id = _cell_text(transaction_cell) or None

            summary_text = re.sub(r"\W+", "", description.casefold())
            if summary_text in SUMMARY_WORDS and booking_date is None:
                warnings.append("possible_summary_row")
                if "footer_or_summary_rows_present" not in parser_warnings:
                    parser_warnings.append("footer_or_summary_rows_present")

            payload: dict[str, str] = {}
            for index, visible_value in enumerate(visible_values):
                if not visible_value:
                    continue
                header = (
                    headers[index]
                    if index < len(headers) and headers[index]
                    else f"column_{index + 1}"
                )
                payload[f"{get_column_letter(index + 1)}:{header}"] = visible_value
            row_number = original_row[0].row
            candidates.append(
                XlsxParsedCandidate(
                    sourceRow=row_number,
                    rawPayload=payload,
                    rawDescription=description,
                    rawBookingDate=booking_raw,
                    rawValueDate=value_raw,
                    rawAmount=raw_amount or raw_debit or raw_credit or None,
                    rawCurrency=raw_currency,
                    rawBalance=raw_balance or None,
                    bookingDate=booking_date,
                    valueDate=value_date,
                    amount=amount,
                    currency=currency,
                    direction=cast(Literal["credit", "debit", "unknown"], direction),
                    balanceAfter=balance,
                    counterparty=counterparty,
                    bankTransactionId=transaction_id,
                    confidence=max(0.0, 1.0 - min(len(set(warnings)), 5) * 0.15),
                    fieldConfidence={
                        "bookingDate": 1.0 if booking_date else 0.0,
                        "description": 1.0 if description else 0.0,
                        "amount": 1.0 if amount else 0.0,
                    },
                    warnings=list(dict.fromkeys(warnings)),
                    rawCells=raw_cells,
                )
            )

        if candidates:
            first_cells = {cell.column - 1: cell for cell in candidates[0].rawCells}
            cell_type_hints = {
                field: first_cells[index].rawType
                for field in MAPPED_FIELDS
                if (index := getattr(mapping, field)) is not None and index in first_cells
            }
            number_format_hints = {
                field: first_cells[index].numberFormat
                for field in MAPPED_FIELDS
                if (index := getattr(mapping, field)) is not None
                and index in first_cells
                and first_cells[index].numberFormat
            }
            mapping = mapping.model_copy(
                update={
                    "cellTypeHints": cell_type_hints,
                    "numberFormatHints": number_format_hints,
                }
            )

        return XlsxParserResult(
            contractVersion="racio.parser.v2",
            source=XlsxParserSource(
                sourceType="xlsx",
                filename=filename,
                mediaType=media_type,
                sheetName=worksheet.title,
                sheetIndex=sheet_index,
                headerRow=mapping.headerRow,
                firstDataRow=mapping.firstDataRow,
                lastDataRow=mapping.lastDataRow,
                workbookDateSystem=inspection.dateSystem,
                formulaCellCount=sheet_info.formulaCellCount,
                mergedRangeCount=sheet_info.mergedRangeCount,
                detectedLanguage=None,
            ),
            mapping=XlsxMappingResult(
                status=cast(Literal["confident", "ambiguous", "invalid"], mapping_status),
                columns=mapping,
                confidence=mapping_confidence,
                warnings=mapping_warnings,
            ),
            candidates=candidates if mapping_status == "confident" else [],
            warnings=list(dict.fromkeys(parser_warnings)),
        )
    finally:
        workbook.close()
        cached_workbook.close()
